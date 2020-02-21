#==========================================================================================
# Creator: Ellis, Kevin
# Organization: n/a
# Description: wip.py What IP allows user to get the network information for a given ip.
# Date: 2020207
# Version: 1.4
# Requirements: A properly formated xlsx file named networks.xlsx with the network information
#===========================================================================================
from netaddr import IPNetwork, IPAddress, valid_ipv4
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import re, shelve, pathlib, os, json, pprint, netaddr
import tkinter as tk
from tkinter import ttk, messagebox

#==========================
# Define Classes
#==========================
class Label_Frame(ttk.LabelFrame):
	"""A custom tkinter LabelFrame class"""
	def __init__(self, window, text, col, row, **args):
		super().__init__(window, text=text, **args)
		self.col = col
		self.row = row
		self.grid(column=self.col, row=self.row)

class Label(ttk.Label):
	"""A custom tkinter Label class"""
	def __init__(self, window, text, col, row, **args):
		super().__init__(window, text=text, **args)
		self.col = col
		self.row = row
		self.grid(column=self.col, row=self.row)

class Button(ttk.Button):
	"""A custom tkinter Button class"""
	def __init__(self, window, text, col, row, **args):
		super().__init__(window, text=text, **args)
		self.col = col
		self.row = row
		self.grid(column=self.col, row=self.row)

class Entry(ttk.Entry):
	"""A custom tkinter Entry class"""
	def __init__(self, window, col, row, **args):
		super().__init__(window, **args)
		self.col = col
		self.row = row
		self.grid(column=self.col, row=self.row)

class Network():
	"""A class to hold each networks information from the excel sheet"""
	def __init__(self, agency=None, ips=None, netname=None, org=None, cities=None):
		self.agency = agency
		self.ips = ips
		self.netname = netname
		self.org = org
		self.cities = cities

#=========================
# Define gloabal variables
#=========================	
datf = os.path.normpath(pathlib.Path.home() / 'wip.dat')
networks = []

#=========================
# start of wip functions
#=========================	
def update_db():
	"""this allows the user to update the database binary file"""
	print(f'database file: <{datf}> doesn\'t exsist\ncreating <{datf}>...')

	#==================================
	# Load the workbook into a wb object
	# declare global variables
	#==================================
	wb = load_workbook('networks.xlsx')
	sheet = wb.active
	num_rows = sheet.max_row
	num_column = sheet.max_column

	#====================================
	# create regex strings for useful data
	#====================================
	ip_regex = re.compile(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})') # matches ipv4 address
	cider_regex = re.compile(r'(\/\d{1,3})') # matches cider
	replace = re.compile(r'(\s\(\*\))')

	#======================================
	# get the ip data from the sheet
	#======================================
	for i in range(3, num_rows): # start at the third row
		agency = (sheet['A' + str(i)].value)
		ips = (sheet['B' + str(i)].value + sheet['C' + str(i)].value)
		ips = replace.sub('', ips) # removes the (*) string from some of the ips
		netname = (sheet['F' + str(i)].value)
		org = (sheet['J' + str(i)].value)
		cities = (sheet['K' + str(i)].value)
		networks.append(Network(agency, ips, netname, org, cities))
	
	#=====================================================================
	# save the data to a shelve file so the program doesn't
	# have to process the data even if there is no updates to the database
	#=====================================================================
	content = shelve.open(datf)
	content['networks'] = networks # write the data to the shelve file
	print('done...')
	return(networks)

def check_db():
	""" check if the wip.dat file exists"""
	if os.path.exists(datf + '.dat'): # if the database exists then load the info
		content = shelve.open(datf)
		networks = content['networks']
		content.close()
	else: # if the path doesn't exist then create the file
		networks = update_db()
	return networks, datf
	
def get_last(network):
	"""get the last usable ip address from network range"""
	return IPAddress(IPNetwork(network).last - 1)

def get_first(network):
	"""get the first usable ip address from network range"""
	return IPAddress(IPNetwork(network).first + 1)

def do_things(entry):
	"""began the parsing the data for the ip and output the network information to the screen"""
	what_ip = entry.get()
	if not netaddr.valid_ipv4(what_ip):
		lv1.set(f'"{what_ip}" is not a valid ip addess. please use a valid ipv4 address')
	else:
		lv1.set(f'No data for <{what_ip}>\nUpdate the <networks.xlsx> database\nThen select <File>, <Update Database>')
	for net in networks:
		if IPAddress(what_ip) in IPNetwork(net.ips):
			output = f'''IP: {what_ip}
AGENCY: {net.agency}
NETWORK: {net.ips}
NETWORK NAME: {net.netname}
ORGANIZATION: {net.org}
CITY: {net.cities}\n'''
			lv1.set(output)

def _quit():
	"""gracefully close the application"""
	win.quit()
	win.destroy()
	quit()
	
def show_about():
	"""Show application creator information"""
	msg = messagebox
	msg.showinfo("", '''Creator: Ellis, Kevin
Organization: n/a
Description: Retrieve the network information from a database
Date: 2020208
Version: 1.4''')
	
def get_network(networkToCheck):
	"""will get the network detail of a given ip with cider"""
	net = networkToCheck.get()
	if netaddr.valid_ipv4(net[:-3]):
		networkInfo = f'''NETWORK: {IPNetwork(net).network}
FIRST HOST: {get_first(net)}
LAST HOST: {get_last(net)}
BROADCAST: {IPNetwork(net).broadcast}
NETMASK: {IPNetwork(net).netmask}
NEXT NETWORK: {IPNetwork(net).next()}\n'''
		networkVar.set(networkInfo)
	else:
		networkVar.set(f'**Error**: "{net}" is not a valid ip\nExample: "192.168.1.0/24"')

#=======================
# End of wip functions
#=======================
def main():
	"""main function that runs the application"""
	version = 'Version: 1.4'
	#==============
	# Create menu
	#==============
	menubar = tk.Menu(win)
	file_menu = tk.Menu(menubar, tearoff=0)
	file_menu.add_command(label='Update Database', command=update_db)
	file_menu.add_separator()
	file_menu.add_command(label='Exit', command=_quit)
	
	help_menu = tk.Menu(menubar, tearoff=0)
	help_menu.add_command(label='About', command=show_about)
	
	menubar.add_cascade(label='File', menu=file_menu)
	menubar.add_cascade(label='Help', menu=help_menu)
	win.config(menu=menubar)
	
	#==================
	# Create Tabs
	#==================
	tabbed = ttk.Notebook(win)
	tab1 = ttk.Frame(tabbed)
	tab2 = ttk.Frame(tabbed)
	
	#===============================
	# create the main label frames
	#===============================
	lf1 = Label_Frame(tab1, version, 0, 0)
	lf2 = Label_Frame(tab2, version, 0, 0)
	tabbed.add(tab1, text='IP Search')
	tabbed.add(tab2, text='Subnet Calc')
	tabbed.pack(expand=0, fill='both')
	#================================
	# create the elements for lf1
	#================================
	l1 = Label(lf1, 'Search IP: ', 0, 0)
	e1 = Entry(lf1, 1, 0)
	e1.focus() # allows the user to start typing as soon as program is run
	l2 = Label(lf1,'', 0, 1, textvar=lv1)
	b1 = Button(lf1, 'Search', 2, 0, command=lambda: do_things(e1))

	#===============================
	# place the elements on lf1
	#===============================
	l1.grid(sticky=tk.W)
	e1.grid(sticky=tk.W)
	l2.grid(columnspan=3, sticky=tk.W)
	b1.grid(sticky=tk.W)
	
	#===============================
	# create the elements for lf2
	#===============================
	networkLabel = Label(lf2, 'IP with CIDR: ', 0, 0)
	networkEntry = Entry(lf2, 1, 0)
	networkButton = Button(lf2, 'Submit', 2, 0, command=lambda: get_network(networkEntry))
	networkLabelData = Label(lf2, '', 0, 1, textvar=networkVar)
	#===============================
	# place the elements on lf2
	#===============================
	networkLabel.grid(sticky=tk.W)
	networkEntry.grid(sticky=tk.W)
	networkButton.grid(sticky=tk.W)
	networkLabelData.grid(sticky=tk.W, columnspan=3)
	
	#===================================
	# place the frame on the main window
	#===================================
	for child in lf1.winfo_children():
			child.grid_configure(padx=5,pady=10)
			
	for child in lf2.winfo_children():
		child.grid_configure(padx=5,pady=10)
	#===============================
	# start the main window loop
	#===============================
	#win.iconbitmap('wip.ico')
	win.mainloop()
	
if __name__=='__main__':
	data, datf = check_db()
	# Create main window
	win = tk.Tk()
	win.title('wip')
	lv1 = tk.StringVar()
	networkVar = tk.StringVar()
	main()
